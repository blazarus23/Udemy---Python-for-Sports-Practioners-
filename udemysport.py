#### UDEMY COURSE - INTRODUCTION TO CODING FOR SPORTS PRACTITIONERS

### COMPARISON OPERATORS
ManCity = 62
Liverpool = 62
Tottenham = 57
"""
# == equals to
print('Does Man City and Liverpool have the same points?')
print(ManCity == Liverpool)
# != does not equal to
print('Do Man City and Tottenham have different points?')
print(ManCity != Tottenham)
# > (greater than)
print('Does Man City have more points that Tottenham?')
print(ManCity > Tottenham)
# >= (greater than or equal to)
print('Does Man City have more points than Liverpool?')
print(ManCity > Liverpool)
# < (less than)
print('Does Liverpool have less points than Tottenham?')
print(Liverpool < Tottenham)
# <= (less than or equal to)
print('Does Liverpool have less points than Man City?')
print(Liverpool <= ManCity)
"""
### LISTS
grades = ['A', 'B', 'C', 'D']
#print(grades)

# call on a specific item in the list
#print(grades[2])

# change an item in the list
grades[3] = 'E'
#print(grades)

# add another item to the end of the list
grades.append('F')
#print(grades)

# insert item into specific location
grades.insert(0, 'S')
#print(grades)

# remove an item
grades.remove('F')
#print(grades)

### FOR LOOPS
# for items means it will be taking on the first item in grades
#for items in grades:
    #print(items)

# determine the number of times a loop can run
#for x in range(10):
    #print(x)
# range can be used 2 ways:
    # range(5) = 0 to 4
    #range(1,5) = 1 to 4
"""
### IF STATEMENTS
if Liverpool >= Tottenham:
    print("Liverpool is Winning")
else:
    print("Tottenham is Winning")
    
### PUTTING IT ALL TOGETHER - OPERATORS, LISTS, FOR LOOPS AND IF STATEMENTS
scores = [41, 30, 25, 22, 50, 60, 40]
grades = ['A', 'B', 'C', 'D', 'E']
extra_training = []

for x in scores:
    if x >= 41:
        print(x)
        print('Grade is equal to ' + grades[0])
    if 36 <= x < 41:
        print(x)
        print('Grade is equal to ' + grades[1])
    if x <= 26:
        print(x)
        print('Grade is equal to ' + grades [4])
        extra_training.append(x)
print(extra_training)
"""

#### WORKING WITH EXCEL AND CREATING ATHLETE REPORTS
from openpyxl import *  #using an asterix imports the entire package.
from mailmerge import MailMerge
from datetime import *
import pandas as pd

pd.options.display.width= None
pd.options.display.max_columns= None
pd.set_option('display.max_rows', 3000)
pd.set_option('display.max_columns', 3000)

# read in data using openpyxl function
wb = load_workbook(filename = "NAPFA Scores.xlsx", data_only=True) # data_only = py will ignore formulas

# to define specific sheets to read
sheet = wb['Sheet1']
#sheet1 = wb['Class6B'] # use this if you want to read in multiple sheets

# find out how many rows in the excel file
maxrow = sheet.max_row
#print(maxrow)

# to access a particular row
# sheet.cell[row,column]
#Sidney_SitUp = sheet.cell(2,5).value
#print(Sidney_SitUp)

# change a cell value then resave the file *optional
#sheet.cell(2,5).value = 70
#wb.save('NAPFA+Scores.xlsx')
#print(Sidney_SitUp)

#### CREATING INDIVIDUALIZED ATHLETE REPORTS
# Use the MailMerge function to call on the template
doc = MailMerge('Athlete+Report+Template.docx')

# Check the merge fields are in the document
#print(doc.get_merge_fields())

# Input our data into the corresponding merge fields
#doc.merge(PNAME = 'Carey Tan', PTEAM = 'Tennis')

# Save the document
#doc.write('Test1.docx')

### USING FOR LOOPS TO CREATE MULTIPLE REPORTS FOR ALL DATA

# Create a for loop to grab all the values we want
for x in range(2, maxrow + 1):
    doc = MailMerge('Athlete+Report+Template.docx')

    pname = sheet.cell(x, 1).value
    pclass = sheet.cell(x, 2).value
    gender = sheet.cell(x, 3).value

    pdob = sheet.cell(x, 4).value
    pdob = datetime.strftime(pdob, '%d/%m/%Y')

    psitups = sheet.cell(x, 5).value
    pbroadjump = sheet.cell(x, 6).value
    psitreach = sheet.cell(x, 7).value
    ppullups = sheet.cell(x, 8).value
    pshuttlerun = sheet.cell(x, 9).value
    pkmrun = sheet.cell(x, 10).value
    pheight = sheet.cell(x, 11).value
    pweight = sheet.cell(x, 12).value
    bmi = sheet.cell(x, 13).value
    result = sheet.cell(x, 15).value
    pteam = sheet.cell(x, 16).value
    pseatedheight = sheet.cell(x, 17).value

    testdate = sheet.cell(x, 14).value
    testdate = datetime.strftime(testdate, '%d/%m/%Y')

    # Merge the fields
    doc.merge(PNAME=pname,
              CLASS=pclass,
              DATE=str(testdate),
              DOB=str(pdob),
              situps=str(psitups),
              broadjump=str(pbroadjump),
              sitreach=str(psitreach),
              pullups=str(ppullups),
              shuttlerun=str(pshuttlerun),
              kmrun=str(pkmrun),
              height = str(pheight),
              weight = str(pweight),
              bmi=str(bmi),
              RESULT=result,
              seated=str(pseatedheight),
              PTEAM=pteam)
    # Save the document
    doc.write('/Users/brendanlazarus/PycharmProjects/UdemyCourse/Output/ ' + pname + '.docx')

### CREATING INTERACTIVE GRAPHS
## PREPARING THE DATA FOR THE INTERACTIVE PLOT
from openpyxl import *
from bokeh.plotting import figure, output_file, show
import warnings
warnings.simplefilter("ignore")

wb = load_workbook(filename = 'Height.xlsx', data_only=True)
sheet = wb['Height']
maxrows = sheet.max_row
#print(maxrows)

xvalues = [2014, 2015, 2016, 2017, 2018, 2019]
yvalues = [] #creates an empty list
"""
# Method 1: use iter_rows that will grab data iterating through each row
# min_col = which column the data will start, min_row = which row the data will start, max_row = where data will stop
for row in sheet.iter_rows(min_col=3, min_row=2, max_row=5, values_only=True):
    #print(row)
    yvalues.append(row)
#print(yvalues)
"""
# Method 2: grab the individual cell data for each participant
for x in range(2,6):
    temp=[]
    data2014 = sheet.cell(x,3).value
    data2015 = sheet.cell(x,4).value
    data2016 = sheet.cell(x,5).value
    data2017 = sheet.cell(x,6).value
    data2018 = sheet.cell(x,7).value
    data2019 = sheet.cell(x,8).value
    temp = [data2014,data2015,data2016,data2017,data2018,data2019]
    yvalues.append(temp)
#print(yvalues)

# create an average for male and females
gender=[]
for gen in sheet.iter_rows(min_row=1, min_col=2, values_only=True):
    gender.append(gen)





# get an average value for each year's data
combinedList = []
for col in sheet.iter_cols(min_col=3, min_row=2, values_only=True):
    combinedList.append(col)

combined2014 = sum(combinedList[0])/len(combinedList[0]) # summing the values in the list divided by the length of values
combined2015 = sum(combinedList[1])/len(combinedList[1])
combined2016 = sum(combinedList[2])/len(combinedList[2])
combined2017 = sum(combinedList[3])/len(combinedList[3])
combined2018 = sum(combinedList[4])/len(combinedList[4])
combined2019 = sum(combinedList[5])/len(combinedList[5])

combined = [combined2014, combined2015, combined2016, combined2017, combined2018, combined2019]
#print(combined)



"""
## USING BOKEH TO CREATE THE INTERACTIVE PLOT
# Output to static HTML file
output_file("heightplot.html")

# Create a new plot with the title and axis label
p = figure(title="Height Analysis", x_axis_label="Year", y_axis_label="Height", width=1200, height=600)

# add a line renderer with legend and line thickness
p.line(xvalues, combined, legend="Combined Average", line_width=3, color="orange")

# to add in the individual participant line use a for loop
for x in range(0,12): # this will look and use values 0 1 2 3
    #p.line(xvalues, yvalues[x], legend='Participant' + str(x))
    #p.line(xvalues, yvalues1[x], legend='Participant' + str(x), color="red")

# To make the legend interactive
p.legend.click_policy="hide"

# show results
#show(p)
"""

#### BUILDING A DATABASE GENERATOR FOR SPORTS DATA
from openpyxl import *

workbook = load_workbook(filename="Udemy (Responses).xlsx", data_only=True)

#print(workbook.sheetnames) #print sheetnames that are inside the excel file

sheet = workbook['Sheet 1']
maxrow = sheet.max_row
namelist = []
for x in range(2, maxrow+1):
    name = sheet.cell(row=x, column=2).value
    namelist.append(name)
#print(namelist)

# remove the duplicates in the list
unique = list(set(namelist))
#print('Number of students: ' + str(len(unique)))
#print(unique)

# create a list with all the student's data that looks like:
    # [[Student 1, 0, 0, 0, 0, 0], [Student 2, 0,0.....]
nostudents = len(unique) # telling us the length (or number) of students in our list
resultslist = [[0,0,0,0,0,0,0] for x in range(nostudents)] #this will create a fake list for the number of students
#print(resultslist)

# Loop through the data and grab what we want





